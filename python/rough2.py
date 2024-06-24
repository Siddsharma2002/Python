import openpyxl
# codefor leven algo
def func(i,j,string1,string2):
    if(i<0) :
        return j+1
    if(j<0) :
        return i+1
    if(string1[i]==string2[j]):
        return func(i-1,j-1,string1,string2)
    else:
       return min(func(i-1,j,string1,string2),func(i,j-1,string1,string2),func(i-1,j-1,string1,string2)) +1
def MainFunc(string1,string2):
    len1=len(string1)
    len2=len(string2)
    x = func(len1-1,len2-1,string1,string2)
    return x
    #print(x)

# code ends for leven
#code for importing xlsx files
workbook=openpyxl.load_workbook("data.xlsx")
workbook2=openpyxl.load_workbook("data2.xlsx")
sheets=workbook.sheetnames
sheets2=workbook2.sheetnames

sheet1=workbook['Sheet1']
sheet2=workbook2['Sheet2']
rows=sheet1.max_row
rows2=sheet2.max_row
columns=sheet1.max_column
columns2=sheet2.max_column
#code for parsing xlsx files
for a in range(1,rows+1):
    for b in range(1,columns+1):
        string1=str(sheet1.cell(a,b).value)
        string2=str(sheet2.cell(a,b).value)
        #Calling  levenstein function
        ans=MainFunc(string1,string2)
        #printing the output
        print(ans)
         


