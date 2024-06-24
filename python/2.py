import openpyxl
workbook=openpyxl.load_workbook("data.xlsx")
sheets=workbook.sheetnames
print(sheets)
sheet1=workbook['Sheet1']
rows=sheet1.max_row
columns=sheet1.max_column
print(f'the number of rows are ',rows)
print(f'the number of columns are ',columns)
for a in range(1,rows+1):
    for b in range(1,columns+1):
        print(sheet1.cell(a,b).value)




