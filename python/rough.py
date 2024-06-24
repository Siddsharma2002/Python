import openpyxl
workbook2=openpyxl.load_workbook("data2.xlsx")
sheets2=workbook2.sheetnames
print(sheets2)
sheet2=workbook2['Sheet2']
rows2=sheet2.max_row
columns2=sheet2.max_column
print(f'the number of rows are ',rows2)
print(f'the number of columns are ',columns2)
for a in range(1,rows2+1):
    for b in range(1,columns2+1):
        print(sheet2.cell(a,b).value) 


for x in range(1,rows2+1):
            for y in range(1,columns2+1):
                print(sheet1.cell(a,b).value," ",sheet2.cell(a,b).value) 


